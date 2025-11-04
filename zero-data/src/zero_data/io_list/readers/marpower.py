from pathlib import Path
from typing import List

from openpyxl import load_workbook
import polars as pl

from ..types import IOValue, IOTopic, IOResult
from ..base import ReaderBase

_DATA_TYPES = {
    "Float": "REAL",
    "Bool": "BOOLEAN",
    "Uint32": "BIGINT",
    "Uint16": "INTEGER",
    # Typo in AMCS IO list R2.14
    "Unit16": "INTEGER",
    "Int32": "INTEGER",
    "Int16": "INTEGER",
    "String": "STRING",
}


class MarpowerReader(ReaderBase):
    def __init__(self):
        self.topic_prefix = "marpower/"
    
    @staticmethod
    def _read_headers(workbook):
        """Determine the header starting from the first non-empty row of the Excel sheet"""
        last_main_header = None
        for col in workbook.columns:
            if col[0].value is None and col[1].value is None:
                break
            elif col[0].value is not None:
                last_main_header = col[0].value

            headers = [
                header
                for header in [last_main_header, col[1].value]
                if header is not None
            ]
            yield " ".join(headers)

    @staticmethod
    def _normalize_marpower_io_list(df: pl.DataFrame):
        """Normalize the IO list by renaming columns and filtering out unnecessary rows"""
        renamed_df = df.rename(lambda c: c.replace(" ", "_").lower())
        filter_df = (
            renamed_df.filter(pl.col("deleted").is_null())
            .filter(pl.col("system") != "SPARE")
            .filter(pl.col("tag") != "SPARE")
        )
        typed_df = (
            filter_df.with_columns(
                pl.col("target_type").replace_strict(_DATA_TYPES).alias("data_type")
            )
            .with_columns(tag=pl.col("tag").str.replace_all(r"-|\.", "_"))
        )
        return typed_df.select([
            "device",
            "tag",
            "yard_tag",
            "target_type",
            "terminal",
            "cabinet",
            "system",
            "description",
            "unit",
            "precision",
            "data_type",
            "mqtt_topic",
            "mqtt_json_path",
        ]).cast({
            "device": pl.String,
            "tag": pl.String,
            "yard_tag": pl.String,
            "target_type": pl.String,
            "terminal": pl.String,
            "cabinet": pl.String,
            "system": pl.String,
            "description": pl.String,
            "unit": pl.String,
            "precision": pl.String,
            "data_type": pl.String,
            "mqtt_topic": pl.String,
            "mqtt_json_path": pl.String
        })

    def _get_io_topics(self, df: pl.DataFrame) -> List[IOTopic]:
        """Get the IO topics from the DataFrame"""
        result = []
        for row in (
            df
            .drop_nulls("mqtt_topic")
            .group_by("mqtt_topic")
            .agg(pl.col("mqtt_json_path"), pl.col("data_type"))
            .iter_rows(named=True)
        ):
            topic = self.determine_topic(row)
            values = [
                IOValue.from_json_path(json_path=jp, data_type=dt)
                for jp, dt in zip(row["mqtt_json_path"], row["data_type"])
            ]
            result.append(IOTopic(topic, values))
        return result

    def determine_topic(self, row: dict) -> str:
        """Create the topic out of a fixed prefix and a name that is a function of the IO list"""
        return self.topic_prefix + row["mqtt_topic"].replace(" ", "-").replace("+", "").lower()

    @classmethod
    def _read_bordered_column(cls, ws, col: int):
        """Read a column from the Excel sheet, returning only the values with borders"""
        last_val = None
        for cell in next(ws.iter_cols(col, col, 3)):
            if cell.border.top is not None and cell.border.top.style is not None:
                last_val = cls.convert_value(cell.value)
            yield last_val

    @classmethod
    def _read_marpower_excel(cls, path: Path) -> pl.DataFrame:
        """Read the AMCS Excel file and return a DataFrame"""
        workbook = load_workbook(path, data_only=True)
        headers = cls._read_headers(workbook["IO-List"])
        data = {
            header: cls._read_bordered_column(workbook["IO-List"], index + 1)
            for index, header in enumerate(headers)
        }
        return pl.DataFrame(data).filter(~pl.all_horizontal(pl.all().is_null()))

    def read_io_list(self, paths: List[Path]) -> IOResult:
        """Read the IO list from the given paths and return an IOResult"""
        io_lists = [
            self._normalize_marpower_io_list(self._read_marpower_excel(path)) for path in paths
        ]
        io_list = pl.concat(io_lists)

        topics = self._get_io_topics(io_list)

        return IOResult(io_list, topics)
