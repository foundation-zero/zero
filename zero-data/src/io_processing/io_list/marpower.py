from pathlib import Path
from typing import List

from openpyxl import load_workbook
from polars import DataFrame
import polars as pl

from io_processing.io_list import IOValue, IOTopic


_DATA_TYPES = {
    "Float": "REAL",
    "Bool": "BOOLEAN",
    "Uint32": "BIGINT",
    "Int32": "INTEGER",
}


def read_headers(ws):
    last_main_header = None
    for col in ws.columns:
        if col[0].value is None and col[1].value is None:
            break
        elif col[0].value is not None:
            last_main_header = col[0].value

        headers = [
            header for header in [last_main_header, col[1].value] if header is not None
        ]
        yield " ".join(headers)


def convert_value(val):
    if val is None:
        return None
    elif isinstance(val, str):
        return val
    elif int(val) == val:
        return str(int(val))
    else:
        return str(val)


def read_bordered_column(ws, col):
    last_val = None
    for cell in next(ws.iter_cols(col, col, 3)):
        if cell.border.top.style is not None:
            last_val = convert_value(cell.value)
        yield last_val


def read_amcs_excel(path) -> DataFrame:
    wb = load_workbook(path, data_only=True)
    headers = read_headers(wb["IO-List"])
    data = {
        header: read_bordered_column(wb["IO-List"], index + 1)
        for index, header in enumerate(headers)
    }
    return DataFrame(data).filter(~pl.all_horizontal(pl.all().is_null()))


def normalize_amcs_io_list(df: DataFrame):
    renamed_df = df.rename(lambda c: c.replace(" ", "_").lower())
    filter_df = (
        renamed_df.filter(pl.col("deleted").is_null())
        .filter(pl.col("system") != "SPARE")
        .filter(pl.col("tag") != "SPARE")
    )
    typed_df = filter_df.with_columns(
        pl.col("target_type").replace_strict(_DATA_TYPES).alias("data_type")
    ).with_columns(yard_tag=pl.col("yard_tag").str.replace_all(r"-|_", "-"))
    return typed_df


def io_topics(df: DataFrame) -> List[IOTopic]:
    result = []
    for row in (
        df.group_by("device", "module")
        .agg(pl.col("tag"), pl.col("data_type"))
        .iter_rows(named=True)
    ):
        module_suffix = f"_{row['module']}" if row["module"] else ""
        topic = f"{row['device']}{module_suffix}"
        values = [
            IOValue(name=t, data_type=dt) for t, dt in zip(row["tag"], row["data_type"])
        ]
        result.append(IOTopic(topic, values))
    return result


def read_amcs_excel_to_io_topics(path: Path) -> List[IOTopic]:
    df = read_amcs_excel(path)
    normalized_df = normalize_amcs_io_list(df)

    return io_topics(normalized_df)
