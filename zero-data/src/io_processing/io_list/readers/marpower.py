from pathlib import Path
from typing import List

from openpyxl import load_workbook
import polars as pl

from ..types import IOValue, IOTopic, IOResult
from ..base import ReaderBase

_DATA_TYPES = {
    "Float": "REAL",
    "Bool": "BOOLEAN",
    "Uint32": "BIGINT",
    "Int32": "INTEGER",
}


class MarpowerReader(ReaderBase):
    @staticmethod
    def _read_headers(workbook):
        last_main_header = None
        for col in workbook.columns:
            if col[0].value is None and col[1].value is None:
                break
            elif col[0].value is not None:
                last_main_header = col[0].value

            headers = [
                header
                for header in [last_main_header, col[1].value]
                if header is not None
            ]
            yield " ".join(headers)

    @staticmethod
    def _normalize_amcs_io_list(df: pl.DataFrame):
        renamed_df = df.rename(lambda c: c.replace(" ", "_").lower())
        filter_df = (
            renamed_df.filter(pl.col("deleted").is_null())
            .filter(pl.col("system") != "SPARE")
            .filter(pl.col("tag") != "SPARE")
        )
        typed_df = filter_df.with_columns(
            pl.col("target_type").replace_strict(_DATA_TYPES).alias("data_type")
        ).with_columns(yard_tag=pl.col("yard_tag").str.replace_all(r"-|_", "-"))
        return typed_df

    @staticmethod
    def _get_io_topics(df: pl.DataFrame) -> List[IOTopic]:
        result = []
        for row in (
            df.group_by("device", "module")
            .agg(pl.col("tag"), pl.col("data_type"))
            .iter_rows(named=True)
        ):
            module_suffix = f"_{row['module']}" if row["module"] else ""
            topic = f"{row['device']}{module_suffix}"
            values = [
                IOValue(name=t, data_type=dt)
                for t, dt in zip(row["tag"], row["data_type"])
            ]
            result.append(IOTopic(topic, values))
        return result

    @classmethod
    def _read_bordered_column(cls, ws, col: int):
        last_val = None
        for cell in next(ws.iter_cols(col, col, 3)):
            if cell.border.top.style is not None:
                last_val = cls.convert_value(cell.value)
            yield last_val

    @classmethod
    def _read_amcs_excel(cls, path: Path) -> pl.DataFrame:
        workbook = load_workbook(path, data_only=True)
        headers = cls._read_headers(workbook["IO-List"])
        data = {
            header: cls._read_bordered_column(workbook["IO-List"], index + 1)
            for index, header in enumerate(headers)
        }
        return pl.DataFrame(data).filter(~pl.all_horizontal(pl.all().is_null()))

    def read_io_list(self, path: Path) -> IOResult:
        df = self._read_amcs_excel(path)

        io_list = self._normalize_amcs_io_list(df)
        topics = self._get_io_topics(io_list)

        return IOResult(io_list, topics)
